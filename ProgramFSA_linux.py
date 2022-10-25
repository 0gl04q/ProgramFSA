import os
import time

from selenium.webdriver import Firefox
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains

from openpyxl import load_workbook


# Функция подключения драйвера
def fire_fox():
    # Подключение опций
    options = Options()
    options.headless = True
    options.add_argument('--window-size=1920,1920')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')

    # Подключение FireFox
    service = Service(executable_path='./geckodriver')
    driver = Firefox(service=service, options=options)

    # Уменьшение масштаба страницы FireFox
    driver.get("about:preferences")
    driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//*[@id='defaultZoom']"))
    ActionChains(driver).click(driver.find_element(By.XPATH, "//*[@value='50']")).perform()
    return main(driver)


# функция работы программы
def main(driver):
    driver.get("https://support.fsa.gov.ru")  # Сайт ФСА

    folder = r'/opt/ProgramFSA/File'  # Объявление папки для работы с файлами

    for file in os.listdir(folder):  # Перебор файлов со сведениями
        name_company = file[:3]
        print('Текущий файл: ', file)

        # Проверка папки для скриншотов + создание
        file_folder = file.replace('.xlsx', '')
        if not os.path.exists(fr'/opt/ProgramFSA/Screenshot/{file_folder}'):
            os.mkdir(fr'/opt/ProgramFSA/Screenshot/{file_folder}')

        # Открытие файла Excel
        wb = load_workbook(folder + "/" + file)
        sheet = wb.active

        # Ожидание окончания загрузки сайта
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//div[@data-t='for_metrology']")))
        except:
            # получение доступа к элементу сайта
            driver.execute_script("arguments[0].click();",
                                  driver.find_element(By.XPATH, "//div[@data-t='for_metrology']"))

        # проход по строкам начиная с 5
        nom_str = 5
        str_enumeration = sheet.cell(row=nom_str, column=1).value
        while str_enumeration is not None:
            # Открытие формы для заполнения данных
            driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//*[@id='metrology-report']"))

            # Выбор формы для поверки
            Select(driver.find_element(By.NAME, "alType")).select_by_index(3)

            # Проверка наименования организации и внесение рег номера аккредитации
            if name_company == 'АТМ' or name_company == 'СПК':
                Select(driver.find_element(By.XPATH,
                                           "//*[@id='metrologyReportForm']/div[2]/div/div[1]/select")).select_by_index(
                    2)
                driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[2]/div/div[1]/div/input").send_keys(
                    "0001." + sheet.cell(row=1, column=1).value)
            else:
                Select(driver.find_element(By.XPATH,
                                           "//*[@id='metrologyReportForm']/div[2]/div/div[1]/select")).select_by_index(
                    0)
                driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[2]/div/div[1]/div/input").send_keys(
                    sheet.cell(row=1, column=1).value)

            # Внесение сведений в форму
            driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[3]/div/div/input").send_keys(
                sheet.cell(row=1, column=2).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[1]/div/div/input").send_keys(
                str(sheet.cell(row=nom_str, column=1).value) + " " + str(sheet.cell(row=nom_str,
                                                                                    column=2).value) + " Зав.№" + str(
                    sheet.cell(row=nom_str, column=3).value))

            str_time = sheet.cell(row=nom_str, column=4).value.split('.')

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[2]/div/div/input").send_keys(
                str_time[2] + '-' + str_time[1] + '-' + str_time[0])

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[3]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=5).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[4]/div/div/input").send_keys(
                str(sheet.cell(row=nom_str, column=1).value) + " " + str(sheet.cell(row=nom_str,
                                                                                    column=2).value) + " Зав.№" + str(
                    sheet.cell(row=nom_str, column=3).value))

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[5]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=7).value + " " + sheet.cell(row=nom_str, column=8).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[1]/input").send_keys(
                sheet.cell(row=nom_str, column=9).value)
            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[2]/input").send_keys(
                sheet.cell(row=nom_str, column=10).value)
            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[3]/input").send_keys(
                sheet.cell(row=nom_str, column=11).value)

            # Создание скриншота
            driver.save_screenshot('/opt/ProgramFSA/Screenshot/' + file_folder + '/' + str(
                nom_str - 4) + ' ' + str(sheet.cell(row=nom_str, column=4).value) + ' ' +
                                   name_company.replace(' ', '') + ".png")

            print('Сохранение скриншота заполненного счетчика, строка:', nom_str - 4)

            # Проверка отправки формы
            check_str = driver.find_element(By.XPATH,
                                            "//*[@id='metrologyReportForm']/div[3]/div/div/input").get_attribute(
                "value")

            i = len(check_str)

            # Цикл переотправки сведений при ошибках сервера
            while i > 0:
                i = len(
                    driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[3]/div/div/input").get_attribute(
                        "value"))
                if i > 0:
                    driver.execute_script("arguments[0].click();",
                                          driver.find_element(By.XPATH, "//*[@id='metrology-report-submit']"))
                    driver.save_screenshot('/opt/ProgramFSA/Screenshot/Скрин_проверки_работы_сайта.png')
                    time.sleep(20)

            # Увеличение счетчика строки для прохода по строкам
            nom_str = nom_str + 1
            str_enumeration = sheet.cell(row=nom_str, column=1).value
            time.sleep(20)

        # Закрытие файла Excel
        wb.close()

    # Закрытие драйвера
    driver.close()


if __name__ == '__main__':
    print('Начало работы программы')
    print('#######################')
    fire_fox()
    print('#######################')
    print('Конец работы программы')
