import sys
import os
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service

from openpyxl import load_workbook


# функция для работы дополнительных файлов после создания exe
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# функция проверки необходимых файлов
def check_folders():  # Проверка на существование необходимых папок

    if not os.path.exists('C:\\ProgramFSA'):  # Основная папка
        os.mkdir('C:\\ProgramFSA')

    if not os.path.exists('C:\\ProgramFSA\\Screenshot'):  # Папка скриншотов
        os.mkdir('C:\\ProgramFSA\\Screenshot')

    if not os.path.exists('C:\\ProgramFSA\\File'):  # Папка для файлов выгрузки
        os.mkdir('C:\\ProgramFSA\\File')

    folders = ['АТМ', 'МС', 'СПК']  # Создание папок со скриншотами для различных организаций

    for folder in folders:
        if not os.path.exists(f'C:\\ProgramFSA\\Screenshot\\{folder}'):
            os.mkdir(f'C:\\ProgramFSA\\Screenshot\\{folder}')


# функция работы программы
def main():
    # Проверка существования основных папок
    check_folders()

    # Подключение Chrome
    service = Service(executable_path="./chromedriver")

    driver = webdriver.Chrome(service=service)

    # Уменьшение масштаба
    driver.get("chrome://settings/")
    driver.execute_script("chrome.settingsPrivate.setDefaultZoom(0.5)")

    driver.get("https://support.fsa.gov.ru")  # Сайт ФСА

    folder = 'C:\\ProgramFSA\\File'  # Объявление папки для работы с файлами

    for file in os.listdir(folder):  # Перебор файлов со сведениями
        name_company = file[:3]
        print(file)
        # Открытие файла Excel
        wb = load_workbook(folder + "\\" + file)
        sheet = wb.active

        # Ожидание окончания загрузки сайта
        try:
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@data-t='for_metrology']"))
                                            )
        except:
            # получение доступа к элементу сайта
            driver.execute_script("arguments[0].click();",
                                  driver.find_element(By.XPATH, "//div[@data-t='for_metrology']"))

        # проход по строкам начиная с 5
        nom_str = 5
        str_enumeration = sheet.cell(row=nom_str, column=1).value
        while str_enumeration is not None:
            # Открытие формы для заполнения данных
            driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//*[@id='metrology-report']"))

            # Выбор формы для поверки
            Select(driver.find_element(By.NAME, "alType")).select_by_index(3)

            # Проверка наименования организации и внесение рег номера аккредитации
            if name_company == "АТМ":
                Select(driver.find_element(By.XPATH,
                                           "//*[@id='metrologyReportForm']/div[2]/div/div[1]/select")).select_by_index(
                    2)
                driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[2]/div/div[1]/div/input").send_keys(
                    sheet.cell(row=1, column=1).value)
            else:
                Select(driver.find_element(By.XPATH,
                                           "//*[@id='metrologyReportForm']/div[2]/div/div[1]/select")).select_by_index(
                    0)
                driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[2]/div/div[1]/div/input").send_keys(
                    sheet.cell(row=1, column=1).value)

            # Внесение сведений в форму
            driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[3]/div/div/input").send_keys(
                sheet.cell(row=1, column=2).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[1]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=1).value + " " + sheet.cell(row=nom_str,
                                                                           column=2).value + " Зав.№" + sheet.cell(
                    row=nom_str, column=3).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[2]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=4).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[3]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=5).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[4]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=1).value + " " + sheet.cell(row=nom_str,
                                                                           column=2).value + " Зав.№" + sheet.cell(
                    row=nom_str, column=3).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[5]/div/div/input").send_keys(
                sheet.cell(row=nom_str, column=7).value + " " + sheet.cell(row=nom_str, column=8).value)

            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[1]/input").send_keys(
                sheet.cell(row=nom_str, column=9).value)
            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[2]/input").send_keys(
                sheet.cell(row=nom_str, column=10).value)
            driver.find_element(By.XPATH, "//*[@id='measurementsForm']/div[6]/div[3]/input").send_keys(
                sheet.cell(row=nom_str, column=11).value)

            # Создание скриншота
            driver.save_screenshot("C:\\ProgramFSA\\Screenshot\\" + name_company.replace(' ', '') + "\\" + str(
                nom_str - 4) + "_" + sheet.cell(row=nom_str, column=4).value + " " + name_company.replace(' ',
                                                                                                          '') + ".png")

            # Проверка отправки формы
            check_str = driver.find_element(By.XPATH,
                                            "//*[@id='metrologyReportForm']/div[3]/div/div/input").get_attribute(
                "value")

            i = len(check_str)

            while i > 0:
                i = len(
                    driver.find_element(By.XPATH, "//*[@id='metrologyReportForm']/div[3]/div/div/input").get_attribute(
                        "value"))
                if i > 0:
                    driver.execute_script("arguments[0].click();",
                                          driver.find_element(By.XPATH, "//*[@id='metrology-report-submit']"))
                    time.sleep(20)

            # Увеличение счетчика строки для прохода по циклу
            nom_str = nom_str + 1
            str_enumeration = sheet.cell(row=nom_str, column=1).value
            print(nom_str - 5)

            time.sleep(20)

        # Закрытие файла Excel
        wb.close()


if __name__ == '__main__':
    main()
